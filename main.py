import pathlib
from pathlib import Path
import pandas as pd
import num2words
from num2words import num2words
import openpyxl
from openpyxl import load_workbook

def open_excelsheet():
    book = openpyxl.open("contracts/contracts.xlsx", read_only=True)
    sheet = book.active #workbook.get_sheet_by_name("sheet1")
    return sheet

def autocomplete(row):
    sheet = open_excelsheet()
    # индекс столбца зависит от номера столбца файла из которого нужно считать данные
    client = sheet[row][0].value
    name = sheet[row][4].value
    unit = sheet[row][5].value
    count = sheet[row][6].value
    price = sheet[row][7].value
    nakl = sheet[row][9].value
    data = sheet[row][15].value
    countp = num2words(count, lang='ru')
    sumn = count*price
    sump = num2words(sumn, lang='ru')

    wb = load_workbook('template.xlsx')
    ws = wb.active 
    ws['L19'] = client
    ws['C24'] = name
    ws['T24'] = unit
    ws['W24'] = count
    ws['AB24'] = count
    ws['W25'] = count
    ws['AB25'] = count
    ws['AF24'] = price
    ws['AP13'] = nakl
    ws['AT13'] = data
    ws['N27'] = countp
    ws['AE27'] = sump

    s = wb.save(f'contracts/nakladnaya_{nakl}.xlsx') 
    return(s)
    

for i in range(10,12):
    autocomplete(i)
 
